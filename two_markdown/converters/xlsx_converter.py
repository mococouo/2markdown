from __future__ import annotations

from pathlib import Path

from two_markdown.context import ConversionContext
from two_markdown.errors import MissingDependencyError
from two_markdown.models import ConvertedDocument
from two_markdown.utils import markdown_table


class XlsxConverter:
    name = "xlsx"
    extensions = {".xlsx"}

    def convert(self, path: Path, context: ConversionContext) -> ConvertedDocument:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise MissingDependencyError("openpyxl", "XLSX conversion") from exc

        workbook = load_workbook(path, data_only=True)
        formulas_wb = None
        try:
            formulas_wb = load_workbook(path, read_only=True, data_only=False)
        except Exception:
            pass

        sections = [f"# {path.stem}"]
        warnings: list[str] = []

        for sheet in workbook.worksheets:
            sections.append(f"\n## {sheet.title}\n")
            formula_sheet = _find_sheet(formulas_wb, sheet.title) if formulas_wb else None
            rows, comments = _sheet_rows(sheet, formula_sheet)
            if rows:
                sections.append(markdown_table(rows))
            else:
                sections.append("_Empty sheet._")
            if comments:
                sections.append("\n**Comments:**\n")
                sections.extend(f"- {comment}" for comment in comments)
            merged = list(getattr(sheet, "merged_cells", None).ranges or []) if hasattr(sheet, "merged_cells") else []
            if merged:
                warnings.append(f"Sheet '{sheet.title}' contains {len(merged)} merged range(s); values shown in the top-left cell only.")

        if not workbook.worksheets:
            warnings.append("Workbook has no visible worksheets.")

        metadata = _workbook_metadata(workbook) if context.options.include_metadata else {}
        workbook.close()
        if formulas_wb is not None:
            try:
                formulas_wb.close()
            except Exception:
                pass

        return ConvertedDocument(
            title=metadata.get("title") or path.stem,
            markdown="\n\n".join(sections),
            converter=self.name,
            attachments=context.attachments,
            warnings=warnings,
            metadata=metadata,
        )


def _find_sheet(workbook: object, title: str) -> object | None:
    try:
        return workbook[title]  # type: ignore[index]
    except Exception:
        return None


def _sheet_rows(values_sheet: object, formula_sheet: object | None) -> tuple[list[list[object]], list[str]]:
    formula_map: dict[tuple[int, int], str] = {}
    if formula_sheet is not None:
        try:
            for row in formula_sheet.iter_rows():  # type: ignore[attr-defined]
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_map[(cell.row, cell.column)] = value
        except Exception:
            formula_map = {}

    rows: list[list[object]] = []
    comments: list[str] = []
    for row in values_sheet.iter_rows():  # type: ignore[attr-defined]
        rendered_row: list[object] = []
        row_has_value = False
        for cell in row:
            value = cell.value
            formula = formula_map.get((cell.row, cell.column))
            if formula and value is None:
                rendered_row.append(formula)
                row_has_value = True
            elif formula and value is not None:
                rendered_row.append(f"{value} `{formula}`")
                row_has_value = True
            else:
                rendered_row.append(value)
                if value is not None and str(value).strip():
                    row_has_value = True
            comment = getattr(cell, "comment", None)
            if comment is not None and getattr(comment, "text", None):
                comments.append(f"{cell.coordinate}: {comment.text.replace(chr(10), ' ')}")
        if row_has_value:
            rows.append(rendered_row)
    return rows, comments


def _workbook_metadata(workbook: object) -> dict[str, str]:
    metadata: dict[str, str] = {}
    props = getattr(workbook, "properties", None)
    if props is None:
        return metadata
    pairs = {
        "author": getattr(props, "creator", None),
        "subject": getattr(props, "subject", None),
        "keywords": getattr(props, "keywords", None),
        "category": getattr(props, "category", None),
        "description": getattr(props, "description", None),
    }
    for key, value in pairs.items():
        if value and str(value).strip():
            metadata[key] = str(value).strip()
    title = getattr(props, "title", None)
    if title and str(title).strip():
        metadata["title"] = str(title).strip()
    created = getattr(props, "created", None)
    if created:
        metadata["created"] = _iso(created)
    modified = getattr(props, "modified", None)
    if modified:
        metadata["modified"] = _iso(modified)
    return metadata


def _iso(value: object) -> str:
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        try:
            return iso()
        except Exception:
            pass
    return str(value)
